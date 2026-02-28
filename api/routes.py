from flask import Blueprint, request, jsonify, send_file
from .pipeline import enrich_patient_data
from backend.multi_disease_engine import predict_risks
from backend.disease_registry import DISEASE_REGISTRY
from backend.drive_storage import save_analysis, get_patient_history, get_all_patient_ids, is_drive_connected
from backend.report_generator import generate_report_pdf
from backend.auth import signup as auth_signup, login as auth_login, get_user_by_id
from backend.auth_middleware import require_auth, require_role
import io

import os
import re
import csv
import tempfile
import pdfplumber
import openpyxl

api_bp = Blueprint('api', __name__, url_prefix='/api')


# ──────────────────────────────────────────────────────────────────────
# Known field aliases — maps various ways a lab report might label a
# value to the exact column name our models expect.
#
# IMPORTANT: Keys MUST be lowercase. The lookup normalises input keys
# to lowercase before matching.
# ──────────────────────────────────────────────────────────────────────
FIELD_ALIASES = {
    # ── Exact model column names (lowercased) ──
    # These ensure that headers like "Pregnancies", "BloodPressure",
    # "DiabetesPedigreeFunction" etc. are matched directly.
    "pregnancies": "Pregnancies", "glucose": "Glucose",
    "bloodpressure": "BloodPressure", "skinthickness": "SkinThickness",
    "insulin": "Insulin", "diabetespedigreefunction": "DiabetesPedigreeFunction",
    "salt_intake": "Salt_Intake", "stress_score": "Stress_Score",
    "bp_history": "BP_History", "sleep_duration": "Sleep_Duration",
    "family_history": "Family_History", "exercise_level": "Exercise_Level",
    "smoking_status": "Smoking_Status",

    # ── General / shared ──
    "age": "Age", "patient age": "Age", "yrs": "Age", "years": "Age",
    "bmi": "BMI", "body mass index": "BMI",
    "sex": "sex", "gender": "sex",

    # ── Blood Pressure aliases ──
    "blood pressure": "BloodPressure", "bp": "BloodPressure",
    "systolic bp": "BloodPressure", "systolic": "BloodPressure",

    # ── Diabetes aliases ──
    "fasting glucose": "Glucose", "blood glucose": "Glucose",
    "fbs": "Glucose", "fasting blood sugar": "Glucose",
    "skin thickness": "SkinThickness", "skinfold": "SkinThickness",
    "diabetes pedigree": "DiabetesPedigreeFunction",
    "pedigree function": "DiabetesPedigreeFunction",
    "diabetes pedigree function": "DiabetesPedigreeFunction",

    # ── Hypertension aliases ──
    "salt intake": "Salt_Intake", "salt": "Salt_Intake",
    "stress score": "Stress_Score", "stress": "Stress_Score",
    "sleep duration": "Sleep_Duration", "sleep": "Sleep_Duration",
    "sleep hours": "Sleep_Duration",
    "bp history": "BP_History", "bp_history": "BP_History",
    "medication": "Medication",
    "family history": "Family_History",
    "exercise level": "Exercise_Level", "exercise": "Exercise_Level",
    "smoking status": "Smoking_Status", "smoking": "Smoking_Status",

    # ── CKD aliases ──
    "specific gravity": "Sg", "sg": "Sg", "sp. gravity": "Sg",
    "albumin": "Al", "al": "Al",
    "sugar": "Su", "su": "Su",
    "rbc": "Rbc", "red blood cells": "Rbc",
    "blood urea": "Bu", "bu": "Bu", "urea": "Bu", "bun": "Bu",
    "serum creatinine": "Sc", "sc": "Sc", "creatinine": "Sc",
    "sodium": "Sod", "sod": "Sod", "na": "Sod",
    "potassium": "Pot", "pot": "Pot", "k": "Pot",
    "hemoglobin": "Hemo", "hemo": "Hemo", "hb": "Hemo", "haemoglobin": "Hemo",
    "wbc": "Wbcc", "wbcc": "Wbcc", "white blood cells": "Wbcc",
    "wbc count": "Wbcc",
    "rbcc": "Rbcc", "rbc count": "Rbcc",
    "hypertension": "Htn", "htn": "Htn",

    # ── Cardio aliases ──
    "chest pain": "cp", "cp": "cp", "chest pain type": "cp",
    "cholesterol": "chol", "chol": "chol", "total cholesterol": "chol",
    "resting ecg": "restecg", "restecg": "restecg", "ecg": "restecg",
    "max heart rate": "thalach", "thalach": "thalach", "max hr": "thalach",
    "exercise angina": "exang", "exang": "exang",
    "oldpeak": "oldpeak", "st depression": "oldpeak",
    "slope": "slope", "st slope": "slope",
    "ca": "ca", "major vessels": "ca",
    "thal": "thal", "thalassemia": "thal",
    "trestbps": "trestbps", "resting bp": "trestbps",
    "resting blood pressure": "trestbps",
}


def _resolve_header(raw_header: str):
    """Resolve a raw column header to a model column name, or None."""
    key = raw_header.strip().lower()
    return FIELD_ALIASES.get(key)


def _extract_kv_from_row(row, extracted: dict):
    if row and len(row) >= 2:
        mapped = _resolve_header(str(row[0] or ""))
        val = str(row[-1] or "").strip()
        if mapped and val:
            extracted[mapped] = val

def _parse_key_value_table(table: list, extracted: dict):
    for row in (table or []):
        _extract_kv_from_row(row, extracted)

def _parse_tabular_data(table: list, resolved: list, extracted: dict):
    for data_row in table[1:]:
        for i, col_name in enumerate(resolved):
            if col_name and i < len(data_row):
                val = str(data_row[i] or "").strip()
                if val and col_name not in extracted:
                    extracted[col_name] = val

def _parse_pdf_table(table: list, extracted: dict):
    if not table or len(table) < 2:
        _parse_key_value_table(table, extracted)
        return

    headers = [str(h or "").strip() for h in table[0]]
    resolved = [_resolve_header(h) for h in headers]

    if any(resolved):
        _parse_tabular_data(table, resolved, extracted)
    else:
        _parse_key_value_table(table, extracted)

def _parse_text_lines(full_text: str, extracted: dict):
    for line in full_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        for sep in [":", "=", "–"]:
            if sep in line:
                parts = line.split(sep, 1)
                val = parts[1].strip()
                val_clean = re.sub(
                    r'\s*(mg/dl|mg/l|mmhg|mm hg|g/dl|meq/l|cells/cumm|million/cmm|%|gm/dl|gm%)\s*$',
                    '', val, flags=re.IGNORECASE
                ).strip()
                mapped = _resolve_header(parts[0])
                if mapped and val_clean and mapped not in extracted:
                    extracted[mapped] = val_clean
                break


def _extract_from_pdf(filepath: str) -> dict:
    """Extract key-value pairs from a PDF medical report."""
    extracted = {}

    with pdfplumber.open(filepath) as pdf:
        full_text = ""
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                _parse_pdf_table(table, extracted)

            text = page.extract_text() or ""
            full_text += text + "\n"

        # --- Text pattern extraction ---
        _parse_text_lines(full_text, extracted)

    return extracted


def _extract_kv_from_row_spreadsheet(row, extracted: dict):
    if row and len(row) >= 2:
        mapped = _resolve_header(str(row[0] or ""))
        val = str(row[1] if row[1] is not None else "").strip()
        if mapped and val:
            extracted[mapped] = val

def _extract_from_tabular_row(data_row: list, resolved: list, extracted: dict):
    for i, col_name in enumerate(resolved):
        if col_name and i < len(data_row):
            val = str(data_row[i] if data_row[i] is not None else "").strip()
            if val and col_name not in extracted:
                extracted[col_name] = val

def _parse_tabular_spreadsheet(rows: list, resolved: list, extracted: dict):
    for data_row in rows[1:]:
        _extract_from_tabular_row(data_row, resolved, extracted)
        if extracted:
            break

def _parse_spreadsheet_rows(rows: list, extracted: dict):
    if not rows:
        return
    
    headers = [str(h or "").strip() for h in rows[0]]
    resolved = [_resolve_header(h) for h in headers]

    if any(resolved):
        _parse_tabular_spreadsheet(rows, resolved, extracted)
    else:
        for row in rows:
            _extract_kv_from_row_spreadsheet(row, extracted)

def _extract_from_excel(filepath: str) -> dict:
    """Extract data from an Excel file."""
    extracted = {}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        _parse_spreadsheet_rows(rows, extracted)

    return extracted


def _extract_from_csv(filepath: str) -> dict:
    """Extract data from a CSV file."""
    extracted = {}

    with open(filepath, newline='', encoding='utf-8-sig') as f:
        reader = list(csv.reader(f))

    _parse_spreadsheet_rows(reader, extracted)

    return extracted


def _parse_numeric(val: str):
    """Try to parse a string as a number, return as-is if categorical."""
    val = val.strip()
    try:
        return float(val)
    except ValueError:
        return val


@api_bp.route('/disease-config', methods=['GET'])
def disease_config():
    """Expose the disease registry to the frontend for dynamic form generation.

    Returns a JSON object mapping each disease name to its required columns
    and any field metadata hints the UI needs (type, options for categoricals).
    """
    # Field metadata — tells the frontend how to render each field
    FIELD_META = {
        # Binary 0/1 fields
        "sex":            {"type": "select", "options": [{"label": "Male", "value": 1}, {"label": "Female", "value": 0}]},
        "Rbc":            {"type": "select", "options": [{"label": "Normal", "value": 1}, {"label": "Abnormal", "value": 0}]},
        "Htn":            {"type": "select", "options": [{"label": "Yes", "value": 1}, {"label": "No", "value": 0}]},
        "exang":          {"type": "select", "options": [{"label": "Yes", "value": 1}, {"label": "No", "value": 0}]},
        "fbs":            {"type": "select", "options": [{"label": ">120 mg/dl", "value": 1}, {"label": "≤120 mg/dl", "value": 0}]},
        # Categorical string fields
        "Smoking_Status": {"type": "select", "options": [{"label": "Non-Smoker", "value": "Non-Smoker"}, {"label": "Smoker", "value": "Smoker"}, {"label": "Former", "value": "Former"}]},
        "BP_History":     {"type": "select", "options": [{"label": "Normal", "value": "Normal"}, {"label": "Elevated", "value": "Elevated"}, {"label": "Hypertension", "value": "Hypertension"}]},
        "Medication":     {"type": "select", "options": [{"label": "None", "value": "None"}, {"label": "Antihypertensives", "value": "Antihypertensives"}, {"label": "Other", "value": "Other"}]},
        "Family_History":  {"type": "select", "options": [{"label": "No", "value": "No"}, {"label": "Yes", "value": "Yes"}]},
        "Exercise_Level":  {"type": "select", "options": [{"label": "Low", "value": "Low"}, {"label": "Moderate", "value": "Moderate"}, {"label": "High", "value": "High"}]},
        # Numeric range selects
        "cp":             {"type": "select", "options": [{"label": "0 — Typical Angina", "value": 0}, {"label": "1 — Atypical Angina", "value": 1}, {"label": "2 — Non-anginal", "value": 2}, {"label": "3 — Asymptomatic", "value": 3}]},
        "restecg":        {"type": "select", "options": [{"label": "0 — Normal", "value": 0}, {"label": "1 — ST-T Abnormality", "value": 1}, {"label": "2 — LV Hypertrophy", "value": 2}]},
        "slope":          {"type": "select", "options": [{"label": "0 — Upsloping", "value": 0}, {"label": "1 — Flat", "value": 1}, {"label": "2 — Downsloping", "value": 2}]},
        "ca":             {"type": "select", "options": [{"label": "0", "value": 0}, {"label": "1", "value": 1}, {"label": "2", "value": 2}, {"label": "3", "value": 3}]},
        "thal":           {"type": "select", "options": [{"label": "1 — Normal", "value": 1}, {"label": "2 — Fixed Defect", "value": 2}, {"label": "3 — Reversible Defect", "value": 3}]},
    }

    config = {}
    for disease_name, disease_cfg in DISEASE_REGISTRY.items():
        fields = []
        for col in disease_cfg["required_columns"]:
            meta = FIELD_META.get(col, {"type": "number"})
            fields.append({
                "name": col,
                "type": meta["type"],
                **({"options": meta["options"]} if "options" in meta else {})
            })
        config[disease_name] = {
            "label": disease_name.upper() if len(disease_name) <= 3 else disease_name.replace("_", " ").title(),
            "fields": fields,
        }

    return jsonify(config), 200

ERR_NOT_JSON = {"error": "Request must be JSON"}

@api_bp.route('/predict', methods=['POST'])
@require_auth
def predict():
    """Main REST prediction endpoint (requires auth)."""
    if not request.is_json:
        return jsonify(ERR_NOT_JSON), 400

    data = request.get_json()

    # Validate payload
    if 'patient_data' not in data or not isinstance(data['patient_data'], list) or len(data['patient_data']) == 0:
        return jsonify({"error": "Missing or empty 'patient_data' array"}), 400

    patient_data_list = data['patient_data']
    selected_diseases = data.get('selected_diseases', 'all')

    try:
        # 1. Pipeline: Enrich sparse frontend data into full feature frames
        df = enrich_patient_data(patient_data_list)
        
        # 2. Orchestration: Run hybrid engine prediction
        results = predict_risks(df, selected_diseases)

        # 3. Auto-save to Drive / local storage (NON-BLOCKING)
        # Use patient_id from auth token if patient role, else from payload
        if request.user.get('role') == 'patient':
            patient_id = request.user.get('patient_id', 'UNKNOWN')
        else:
            patient_id = data.get('patient_id', 'UNKNOWN')
        import threading
        def _bg_save():
            try:
                save_analysis(
                    patient_id=patient_id,
                    results=results,
                    input_data=patient_data_list[0] if patient_data_list else {},
                )
            except Exception:
                pass  # Never let storage failure affect anything
        threading.Thread(target=_bg_save, daemon=True).start()

        return jsonify({**results, "__storage": {"stored": True, "backend": "async"}}), 200

    except Exception as e:
        # Return 500, never crash the Flask worker
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500

@api_bp.route('/extract-report', methods=['POST'])
def extract_report():
    """Accept a PDF or Excel report file, extract medical data, return structured JSON.

    Returns:
        { "extracted_data": { "Age": 45, "Glucose": 150, ... }, "fields_found": 8 }
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded. Send a file with key 'file'."}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.pdf', '.xlsx', '.xls', '.csv'):
        return jsonify({"error": f"Unsupported file type '{ext}'. Upload a PDF, Excel, or CSV file."}), 400

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    try:
        f.save(tmp.name)
        tmp.close()

        if ext == '.pdf':
            raw = _extract_from_pdf(tmp.name)
        elif ext == '.csv':
            raw = _extract_from_csv(tmp.name)
        else:
            raw = _extract_from_excel(tmp.name)

        # Parse numeric values
        parsed = {}
        for k, v in raw.items():
            parsed[k] = _parse_numeric(str(v))

        return jsonify({
            "extracted_data": parsed,
            "fields_found": len(parsed),
            "source_file": f.filename,
        }), 200

    except Exception as e:
        return jsonify({"error": "Failed to parse report", "details": str(e)}), 500

    finally:
        os.unlink(tmp.name)


@api_bp.route('/export-pdf', methods=['POST'])
def export_pdf():
    """Generate and return a downloadable PDF report from analysis results.

    Expects JSON body:
    {
        "patient_id": "PAT-1001",
        "results": { ... disease results ... },
        "input_data": { ... optional input used ... }
    }
    """
    if not request.is_json:
        return jsonify(ERR_NOT_JSON), 400

    data = request.get_json()
    patient_id = data.get("patient_id", "UNKNOWN")
    results = data.get("results", {})
    input_data = data.get("input_data", {})

    try:
        pdf_bytes = generate_report_pdf(
            patient_id=patient_id,
            results=results,
            input_data=input_data,
        )
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)

        filename = f"ChronoCare_Report_{patient_id}.pdf"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": "PDF generation failed", "details": str(e)}), 500


@api_bp.route('/timeline/<patient_id>', methods=['GET'])
@require_auth
def timeline(patient_id):
    """Return historical analysis records for a patient.

    Used by the Timeline Intelligence page to build trend graphs.
    Returns both raw records and a flattened history array for charting.
    Patients can only access their own timeline. Doctors can access any.
    """
    # Role-based access control
    user = request.user
    if user['role'] == 'patient' and user.get('patient_id') != patient_id:
        return jsonify({"error": "Access denied. You can only view your own timeline."}), 403

    records = get_patient_history(patient_id)

    # Build flattened history for charts: [{date, diabetes, ckd, ...}, ...]
    history = []
    for record in records:
        ts = record.get("timestamp", "")
        date_str = ts[:10] if len(ts) >= 10 else ts
        entry = {"date": date_str}
        results = record.get("results", {})
        for disease_name, disease_data in results.items():
            if isinstance(disease_data, dict) and "probability" in disease_data:
                entry[disease_name] = disease_data["probability"]
        if len(entry) > 1:  # Has at least one disease
            history.append(entry)

    return jsonify({
        "patient_id": patient_id,
        "records": records,
        "history": history,
        "count": len(records),
    }), 200


@api_bp.route('/patients', methods=['GET'])
@require_auth
@require_role('doctor')
def list_patients():
    """Return all patient IDs that have stored history. Doctor-only."""
    ids = get_all_patient_ids()
    return jsonify({"patient_ids": ids, "count": len(ids)}), 200


@api_bp.route('/storage-status', methods=['GET'])
def storage_status():
    """Check if Google Drive is connected."""
    connected = is_drive_connected()
    return jsonify({
        "google_drive_connected": connected,
        "fallback": "local" if not connected else None,
    }), 200


# ══════════════════════════════════════════════════════════════════════
#  AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════
@api_bp.route('/auth/signup', methods=['POST'])
def auth_signup_route():
    """Create a new user account."""
    if not request.is_json:
        return jsonify(ERR_NOT_JSON), 400

    data = request.get_json()
    try:
        auth_signup(
            name=data.get('name', ''),
            email=data.get('email', ''),
            password=data.get('password', ''),
            role=data.get('role', 'patient'),
        )
        # Auto-login after signup
        login_result = auth_login(data['email'], data['password'])
        return jsonify(login_result), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@api_bp.route('/auth/login', methods=['POST'])
def auth_login_route():
    """Authenticate and return JWT token."""
    if not request.is_json:
        return jsonify(ERR_NOT_JSON), 400

    data = request.get_json()
    try:
        result = auth_login(
            email=data.get('email', ''),
            password=data.get('password', ''),
        )
        return jsonify(result), 200
    except ValueError as e:
        return jsonify({"error": str(e)}), 401


@api_bp.route('/auth/me', methods=['GET'])
@require_auth
def auth_me():
    """Get current authenticated user profile."""
    user = get_user_by_id(request.user['user_id'])
    if not user:
        return jsonify({"error": "User not found"}), 404
    return jsonify(user), 200
